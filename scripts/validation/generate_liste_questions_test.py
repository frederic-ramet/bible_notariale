#!/usr/bin/env python3
"""
Génère la liste simple des questions à tester pour Phase 3.
Lit le dataset validé (ou le dataset original si pas encore validé)
et génère un Excel simple pour la session de tests.
"""

import json
import sys
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def load_questions(dataset_path: Path) -> List[Dict[str, Any]]:
    """
    Charge les questions depuis le dataset JSON.

    Args:
        dataset_path: Chemin vers le fichier dataset JSON

    Returns:
        Liste des questions
    """
    print(f"📂 Chargement des questions : {dataset_path.name}")

    try:
        with open(dataset_path, 'r', encoding='utf-8') as f:
            dataset = json.load(f)

        # Si c'est le dataset validé (Phase 2), il peut avoir une structure différente
        if 'qa_pairs' in dataset:
            questions = dataset['qa_pairs']
        elif isinstance(dataset, list):
            # Si c'est directement une liste de questions
            questions = dataset
        else:
            print(f"❌ Format de dataset non reconnu")
            sys.exit(1)

        print(f"✅ {len(questions)} questions chargées")
        return questions
    except Exception as e:
        print(f"❌ Erreur lors du chargement : {e}")
        sys.exit(1)


def generate_excel(
    questions: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path
):
    """
    Génère le fichier Excel simple avec la liste des questions à tester.

    Args:
        questions: Liste des questions
        template_path: Chemin vers le template Excel
        output_path: Chemin du fichier Excel à créer
    """
    print(f"\n📊 Génération du fichier Excel...")

    # Charger le template
    if not template_path.exists():
        print(f"❌ Template non trouvé : {template_path}")
        print(f"   Exécutez d'abord : python scripts/validation/create_template_liste_questions_test.py")
        sys.exit(1)

    wb = load_workbook(template_path)
    ws = wb["Questions_a_Tester"]

    # Remplir les lignes (à partir de la ligne 4, après titre et instructions)
    for idx, q in enumerate(questions, start=4):
        row = idx

        # Colonne A : Numero (déjà pré-rempli dans le template, mais on peut l'écraser)
        ws[f'A{row}'] = idx - 3  # Numéro commence à 1

        # Colonne B : Question
        question_text = q.get('question', '')
        ws[f'B{row}'] = question_text
        ws[f'B{row}'].alignment = Alignment(vertical='top', wrap_text=True)

        # Colonne C : Categorie
        categorie = q.get('categorie', '').title()
        if not categorie:
            # Essayer de déduire la catégorie si elle n'est pas présente
            if q.get('necessite_multi_documents', False):
                categorie = 'Multi-documents'
            else:
                categorie = 'Autre'

        ws[f'C{row}'] = categorie
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='top')

        # Colonnes D et E : Testee et Notes_Rapides - laissées vides (à remplir pendant les tests)

    # Sauvegarder
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Fichier créé : {output_path}")
    print(f"\n📋 PROCHAINES ÉTAPES :")
    print(f"   1. Vérifier que le chatbot est accessible")
    print(f"   2. Vérifier que le système tribunal fonctionne")
    print(f"   3. Imprimer ou partager le fichier Excel avec les experts")
    print(f"   4. Organiser la session de tests (1h30, 2-3 experts)")
    print(f"   5. Les experts testent et donnent leur feedback via le système tribunal")


def main():
    """Point d'entrée principal."""
    print("=" * 70)
    print("GÉNÉRATION LISTE DES QUESTIONS À TESTER - PHASE 3")
    print("=" * 70)
    print()

    # Chemins
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    # Essayer d'abord le dataset validé Phase 2, sinon le dataset validé Phase 2 Excel, sinon le dataset original
    dataset_validated_path = project_root / "tests" / "datasets" / "dataset_test_final_20questions.json"
    dataset_excel_path = project_root / "output" / "validation_dataset_20questions.xlsx"
    dataset_original_path = project_root / "tests" / "datasets" / "chatbot_test_dataset.json"

    # Déterminer quel dataset utiliser
    if dataset_validated_path.exists():
        print(f"✅ Utilisation du dataset validé Phase 2 : {dataset_validated_path.name}")
        dataset_path = dataset_validated_path
    elif dataset_excel_path.exists():
        print(f"⚠️  Dataset validé JSON non trouvé")
        print(f"   Utilisation du fichier Excel Phase 2 : {dataset_excel_path.name}")
        print(f"   Note : Il faudra d'abord exécuter integrate_validated_dataset.py")
        print(f"   Pour l'instant, on utilise le dataset original...")
        dataset_path = dataset_original_path
    else:
        print(f"⚠️  Aucun dataset validé trouvé")
        print(f"   Utilisation du dataset original : {dataset_original_path.name}")
        dataset_path = dataset_original_path

    template_path = project_root / "templates" / "liste_questions_a_tester_TEMPLATE.xlsx"
    output_path = project_root / "output" / "liste_questions_a_tester.xlsx"

    # Vérifications
    if not dataset_path.exists():
        print(f"❌ Aucun fichier dataset trouvé")
        sys.exit(1)

    # Charger les questions
    questions = load_questions(dataset_path)

    # Limiter à 20 questions max
    if len(questions) > 20:
        print(f"ℹ️  {len(questions)} questions trouvées, on prend les 20 premières")
        questions = questions[:20]

    # Générer l'Excel
    generate_excel(questions, template_path, output_path)

    print()
    print("=" * 70)
    print("✅ GÉNÉRATION TERMINÉE")
    print("=" * 70)
    print()
    print("💡 RAPPEL : Ce fichier est pour la session de tests (Phase 3)")
    print("   Les experts doivent :")
    print("   1. Tester chaque question dans le chatbot")
    print("   2. Donner leur feedback via le système tribunal")
    print("   3. Cocher 'Testée' dans l'Excel")
    print("   4. Ajouter des notes rapides si nécessaire")


if __name__ == "__main__":
    main()
