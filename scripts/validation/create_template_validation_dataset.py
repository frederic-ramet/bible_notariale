#!/usr/bin/env python3
"""
Crée le template Excel pour la validation du dataset (20 questions).
Ce fichier sera utilisé par l'expert métier lors de la session de validation Phase 2.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import sys
from pathlib import Path


def create_dataset_validation_template(output_path: str = "templates/validation_dataset_20questions_TEMPLATE.xlsx"):
    """
    Crée un template Excel pour la validation du dataset de questions.

    Args:
        output_path: Chemin du fichier Excel à créer
    """
    # Créer le classeur
    wb = Workbook()

    # ===== ONGLET 1 : INSTRUCTIONS =====
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # Titre
    ws_instructions['A1'] = "GUIDE DE VALIDATION DU DATASET DE QUESTIONS"
    ws_instructions['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_instructions['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_instructions.merge_cells('A1:E1')
    ws_instructions.row_dimensions[1].height = 30

    # Contenu des instructions
    instructions = [
        ("", ""),
        ("OBJECTIF", "Valider que les 20 questions de test sont réalistes et juridiquement exactes"),
        ("DURÉE", "1h30 (environ 3-4 minutes par question)"),
        ("", ""),
        ("COMMENT UTILISER CE FICHIER", ""),
        ("", ""),
        ("1. Pour chaque question :", "Lire la question à voix haute"),
        ("2. Valider si RÉALISTE", "Choisir : Oui / Non / A reformuler"),
        ("", "Si 'A reformuler' : Proposer une meilleure formulation"),
        ("3. Valider les SOURCES", "Les documents proposés permettent-ils de répondre ?"),
        ("", "Choisir : Oui / Non / Incomplet"),
        ("", "Si 'Incomplet' : Indiquer les documents manquants"),
        ("4. Valider les ÉLÉMENTS CLÉS", "Les éléments de réponse sont-ils complets ?"),
        ("", "Choisir : Oui / Incomplet / Incorrect"),
        ("5. Valider la RÉPONSE ATTENDUE", "La réponse est-elle juridiquement exacte ?"),
        ("", "Choisir : Oui / Non / A préciser"),
        ("", "Si 'A préciser' : Ajouter les précisions nécessaires"),
        ("6. Commentaires", "Remarques libres"),
        ("", ""),
        ("IMPORTANT", ""),
        ("", "La validation de la RÉPONSE ATTENDUE est CRITIQUE."),
        ("", "Toute inexactitude juridique doit être corrigée."),
        ("", ""),
        ("RÉPARTITION DES 20 QUESTIONS", ""),
        ("Déontologie", "8 questions (3 faciles, 3 moyennes, 2 pointues)"),
        ("Juridique CCN/RH", "5 questions (2 faciles, 2 moyennes, 1 pointue)"),
        ("Multi-documents", "4 questions (nécessitent plusieurs documents)"),
        ("Edge cases", "3 questions (cas limites ou hors périmètre)"),
    ]

    row = 2
    for col1, col2 in instructions:
        if col1.isupper() and col1 and col2:  # Titres de section
            ws_instructions[f'A{row}'] = col1
            ws_instructions[f'A{row}'].font = Font(bold=True, size=12)
            ws_instructions[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws_instructions[f'B{row}'] = col2
            ws_instructions[f'B{row}'].font = Font(bold=True)
        else:
            ws_instructions[f'A{row}'] = col1
            ws_instructions[f'B{row}'] = col2
        row += 1

    # Ajuster les largeurs de colonnes
    ws_instructions.column_dimensions['A'].width = 35
    ws_instructions.column_dimensions['B'].width = 70

    # ===== ONGLET 2 : VALIDATION_QUESTIONS =====
    ws_validation = wb.create_sheet("Validation_Questions")

    # En-têtes
    headers = [
        "ID",
        "Question",
        "Categorie",
        "Difficulte",
        "Documents_Sources_Proposes",
        "Elements_Cles_Reponse",
        "Reponse_Attendue_Resumee",
        "Validation_Question",
        "Correction_Question",
        "Validation_Sources",
        "Correction_Sources",
        "Validation_Elements_Cles",
        "Correction_Elements_Cles",
        "Validation_Reponse_Attendue",
        "Correction_Reponse_Attendue",
        "Commentaires"
    ]

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws_validation.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Hauteur de la ligne d'en-tête
    ws_validation.row_dimensions[1].height = 50

    # Largeurs de colonnes
    column_widths = {
        'A': 10,  # ID
        'B': 50,  # Question
        'C': 15,  # Categorie
        'D': 12,  # Difficulte
        'E': 40,  # Documents_Sources_Proposes
        'F': 50,  # Elements_Cles_Reponse
        'G': 50,  # Reponse_Attendue_Resumee
        'H': 18,  # Validation_Question
        'I': 40,  # Correction_Question
        'J': 18,  # Validation_Sources
        'K': 35,  # Correction_Sources
        'L': 20,  # Validation_Elements_Cles
        'M': 40,  # Correction_Elements_Cles
        'N': 22,  # Validation_Reponse_Attendue
        'O': 45,  # Correction_Reponse_Attendue
        'P': 50,  # Commentaires
    }

    for col, width in column_widths.items():
        ws_validation.column_dimensions[col].width = width

    # Ajouter 20 lignes vides pour les données
    for row_idx in range(2, 22):  # Lignes 2 à 21 (20 questions)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_validation.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Hauteur de ligne plus grande pour le contenu
        ws_validation.row_dimensions[row_idx].height = 60

    # ===== LISTES DÉROULANTES =====

    # Validation_Question (colonne H)
    dv_question = DataValidation(
        type="list",
        formula1='"Oui,Non,A reformuler"',
        allow_blank=True
    )
    dv_question.error = 'Veuillez choisir une valeur dans la liste'
    dv_question.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_question)
    dv_question.add(f'H2:H21')

    # Validation_Sources (colonne J)
    dv_sources = DataValidation(
        type="list",
        formula1='"Oui,Non,Incomplet"',
        allow_blank=True
    )
    dv_sources.error = 'Veuillez choisir une valeur dans la liste'
    dv_sources.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_sources)
    dv_sources.add(f'J2:J21')

    # Validation_Elements_Cles (colonne L)
    dv_elements = DataValidation(
        type="list",
        formula1='"Oui,Incomplet,Incorrect"',
        allow_blank=True
    )
    dv_elements.error = 'Veuillez choisir une valeur dans la liste'
    dv_elements.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_elements)
    dv_elements.add(f'L2:L21')

    # Validation_Reponse_Attendue (colonne N)
    dv_reponse = DataValidation(
        type="list",
        formula1='"Oui,Non,A preciser"',
        allow_blank=True
    )
    dv_reponse.error = 'Veuillez choisir une valeur dans la liste'
    dv_reponse.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_reponse)
    dv_reponse.add(f'N2:N21')

    # Mise en forme conditionnelle (colonnes de validation en jaune pâle)
    validation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx in range(2, 22):
        ws_validation[f'H{row_idx}'].fill = validation_fill
        ws_validation[f'J{row_idx}'].fill = validation_fill
        ws_validation[f'L{row_idx}'].fill = validation_fill
        ws_validation[f'N{row_idx}'].fill = validation_fill

    # Mise en forme spéciale pour la colonne Validation_Reponse_Attendue (plus critique)
    critical_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    for row_idx in range(2, 22):
        ws_validation[f'N{row_idx}'].fill = critical_fill
        ws_validation[f'O{row_idx}'].fill = critical_fill

    # Figer les volets (première ligne et 2 premières colonnes)
    ws_validation.freeze_panes = 'C2'

    # Sauvegarder le fichier
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)
    print(f"✅ Template créé : {output_file}")
    print(f"   - Onglet 'Instructions' : Guide complet pour l'expert")
    print(f"   - Onglet 'Validation_Questions' : 20 lignes prêtes à remplir")
    print(f"   - Listes déroulantes configurées pour toutes les validations")
    print(f"   - Colonne 'Validation_Reponse_Attendue' mise en évidence (critique)")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "templates/validation_dataset_20questions_TEMPLATE.xlsx"
    create_dataset_validation_template(output_path)
