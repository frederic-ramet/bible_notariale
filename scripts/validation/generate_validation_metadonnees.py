#!/usr/bin/env python3
"""
Génère le fichier Excel de validation des métadonnées pour Phase 1.
Lit les fichiers .metadata.json, sélectionne 20 documents selon les critères,
et génère un Excel pré-rempli prêt pour la session de validation.
"""

import json
import sys
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def load_metadata_files(metadata_dir: Path) -> List[Dict[str, Any]]:
    """
    Charge tous les fichiers .metadata.json du répertoire.

    Args:
        metadata_dir: Chemin vers le répertoire des métadonnées

    Returns:
        Liste de dictionnaires contenant les métadonnées
    """
    metadata_files = list(metadata_dir.glob("*.metadata.json"))
    print(f"📂 Trouvé {len(metadata_files)} fichiers de métadonnées")

    metadatas = []
    for filepath in metadata_files:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                data['_filepath'] = str(filepath)  # Garder le chemin pour référence
                metadatas.append(data)
        except Exception as e:
            print(f"⚠️  Erreur lecture {filepath.name}: {e}")

    print(f"✅ {len(metadatas)} fichiers chargés avec succès")
    return metadatas


def select_20_documents(metadatas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Sélectionne 20 documents selon les critères de la méthodologie.

    Critères :
    - 10 documents avec priorité 10 (si le champ existe)
    - 5 documents représentatifs par type
    - 5 documents avec peu de mots-clés (potentiellement problématiques)

    Args:
        metadatas: Liste de toutes les métadonnées

    Returns:
        Liste de 20 métadonnées sélectionnées
    """
    print("\n🔍 Sélection des 20 documents...")

    selected = []

    # Critère 1 : Documents avec priorité 10 (si disponible)
    # Note : Le champ priorité n'existe pas encore, on le simulera avec des docs critiques
    critical_keywords = ['rpn', 'code', 'deontologie', 'circulaire_01_25', 'guide_negociation']
    priority_10_docs = [
        m for m in metadatas
        if any(kw in m.get('document_id', '').lower() for kw in critical_keywords)
    ]

    # Prendre les 10 premiers (ou moins si pas assez)
    for doc in priority_10_docs[:10]:
        selected.append(doc)
        print(f"  ✓ Priorité 10 : {doc.get('nom_fichier', 'N/A')}")

    # Critère 2 : 5 documents représentatifs par type
    remaining = [m for m in metadatas if m not in selected]

    # Grouper par type de document
    by_type = {}
    for m in remaining:
        doc_type = m.get('classification', {}).get('type_document', 'autre')
        if doc_type not in by_type:
            by_type[doc_type] = []
        by_type[doc_type].append(m)

    # Prendre 1 document de chaque type (jusqu'à 5)
    types_selected = []
    for doc_type, docs in sorted(by_type.items()):
        if len(types_selected) >= 5:
            break
        if docs:
            selected.append(docs[0])
            types_selected.append(doc_type)
            print(f"  ✓ Type '{doc_type}' : {docs[0].get('nom_fichier', 'N/A')}")

    # Critère 3 : 5 documents avec peu de mots-clés (potentiellement problématiques)
    remaining = [m for m in metadatas if m not in selected]

    # Trier par nombre de mots-clés (croissant)
    remaining_sorted = sorted(
        remaining,
        key=lambda m: len(m.get('mots_cles', []))
    )

    for doc in remaining_sorted[:5]:
        selected.append(doc)
        nb_keywords = len(doc.get('mots_cles', []))
        print(f"  ✓ Peu de mots-clés ({nb_keywords}) : {doc.get('nom_fichier', 'N/A')}")

    # Si on n'a pas 20 documents, compléter avec des docs aléatoires
    if len(selected) < 20:
        remaining = [m for m in metadatas if m not in selected]
        for doc in remaining[:20 - len(selected)]:
            selected.append(doc)
            print(f"  ✓ Complément : {doc.get('nom_fichier', 'N/A')}")

    print(f"\n✅ {len(selected)} documents sélectionnés")
    return selected[:20]  # S'assurer qu'on a exactement 20


def generate_excel(
    selected_docs: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path
):
    """
    Génère le fichier Excel pré-rempli avec les 20 documents sélectionnés.

    Args:
        selected_docs: Liste des 20 documents sélectionnés
        template_path: Chemin vers le template Excel
        output_path: Chemin du fichier Excel à créer
    """
    print(f"\n📊 Génération du fichier Excel...")

    # Charger le template
    if not template_path.exists():
        print(f"❌ Template non trouvé : {template_path}")
        print(f"   Exécutez d'abord : python scripts/validation/create_template_validation_metadonnees.py")
        sys.exit(1)

    wb = load_workbook(template_path)
    ws = wb["Validation_Metadonnees"]

    # Remplir les lignes (à partir de la ligne 2, la ligne 1 est l'en-tête)
    for idx, doc in enumerate(selected_docs, start=2):
        classification = doc.get('classification', {})

        # Colonne A : ID
        ws[f'A{idx}'] = doc.get('document_id', '')

        # Colonne B : Nom_Fichier
        ws[f'B{idx}'] = doc.get('nom_fichier', '')

        # Colonne C : Type_Propose
        doc_type = classification.get('type_document', '')
        label = classification.get('label', '')
        ws[f'C{idx}'] = label if label else doc_type

        # Colonne D : Categories_Proposees
        categories = classification.get('categories_metier', [])
        ws[f'D{idx}'] = ', '.join(categories) if categories else ''

        # Colonne E : Priorite_Proposee (simulation - à ajuster)
        # Pour l'instant, on met 10 pour les docs critiques, 5 par défaut
        doc_id_lower = doc.get('document_id', '').lower()
        if any(kw in doc_id_lower for kw in ['rpn', 'code', 'deontologie', 'circulaire']):
            ws[f'E{idx}'] = 10
        else:
            ws[f'E{idx}'] = 5

        # Colonne F : Mots_Cles_Proposes
        mots_cles = doc.get('mots_cles', [])
        ws[f'F{idx}'] = ', '.join(mots_cles) if mots_cles else ''

        # Colonnes G à M : Laisser vides (à remplir par l'expert)
        # G: Validation_Type, H: Correction_Type, etc.

        # Ajuster l'alignement
        for col in ['A', 'B', 'C', 'D', 'F']:
            ws[f'{col}{idx}'].alignment = Alignment(vertical='top', wrap_text=True)

        ws[f'E{idx}'].alignment = Alignment(horizontal='center', vertical='top')

    # Sauvegarder
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Fichier créé : {output_path}")
    print(f"\n📋 PROCHAINES ÉTAPES :")
    print(f"   1. Ouvrir le fichier Excel : {output_path}")
    print(f"   2. Préparer les PDFs des documents pour consultation")
    print(f"   3. Organiser la session de validation avec l'expert (2h)")
    print(f"   4. Sauvegarder le fichier validé sous : validation_metadonnees_20docs_VALIDEE.xlsx")


def main():
    """Point d'entrée principal."""
    print("=" * 70)
    print("GÉNÉRATION FICHIER VALIDATION MÉTADONNÉES - PHASE 1")
    print("=" * 70)
    print()

    # Chemins
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    metadata_dir = project_root / "_metadata" / "documents"
    template_path = project_root / "templates" / "validation_metadonnees_20docs_TEMPLATE.xlsx"
    output_path = project_root / "output" / "validation_metadonnees_20docs.xlsx"

    # Vérifications
    if not metadata_dir.exists():
        print(f"❌ Dossier métadonnées introuvable : {metadata_dir}")
        sys.exit(1)

    # Charger les métadonnées
    metadatas = load_metadata_files(metadata_dir)

    if len(metadatas) < 20:
        print(f"⚠️  Seulement {len(metadatas)} documents trouvés, moins de 20 requis")
        print(f"   On continuera avec ce qu'on a...")

    # Sélectionner 20 documents
    selected = select_20_documents(metadatas)

    # Générer l'Excel
    generate_excel(selected, template_path, output_path)

    print()
    print("=" * 70)
    print("✅ GÉNÉRATION TERMINÉE")
    print("=" * 70)


if __name__ == "__main__":
    main()
