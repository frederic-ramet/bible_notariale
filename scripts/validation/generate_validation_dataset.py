#!/usr/bin/env python3
"""
Génère le fichier Excel de validation du dataset pour Phase 2.
Lit le fichier chatbot_test_dataset.json, sélectionne 20 questions selon les critères,
et génère un Excel pré-rempli prêt pour la session de validation.
"""

import json
import sys
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def load_dataset(dataset_path: Path) -> Dict[str, Any]:
    """
    Charge le fichier dataset JSON.

    Args:
        dataset_path: Chemin vers le fichier dataset JSON

    Returns:
        Dictionnaire contenant le dataset
    """
    print(f"📂 Chargement du dataset : {dataset_path.name}")

    try:
        with open(dataset_path, 'r', encoding='utf-8') as f:
            dataset = json.load(f)

        total_questions = dataset.get('total_questions', 0)
        print(f"✅ Dataset chargé : {total_questions} questions au total")
        return dataset
    except Exception as e:
        print(f"❌ Erreur lors du chargement du dataset : {e}")
        sys.exit(1)


def select_20_questions(dataset: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Sélectionne 20 questions selon la répartition définie dans la méthodologie.

    Répartition :
    - 8 Déontologie (3 facile, 3 moyen, 2 pointu)
    - 5 Juridique CCN/RH (2 facile, 2 moyen, 1 pointu)
    - 4 Multi-documents (1 facile, 2 moyen, 1 pointu)
    - 3 Edge cases (1 facile, 1 moyen, 1 pointu)

    Args:
        dataset: Dictionnaire contenant toutes les questions

    Returns:
        Liste de 20 questions sélectionnées
    """
    print("\n🔍 Sélection des 20 questions selon la répartition...")

    qa_pairs = dataset.get('qa_pairs', [])
    selected = []

    # Grouper les questions par catégorie et difficulté
    by_cat_diff = {}
    for q in qa_pairs:
        cat = q.get('categorie', 'autre')
        diff = q.get('difficulte', 'moyen')
        key = f"{cat}_{diff}"

        if key not in by_cat_diff:
            by_cat_diff[key] = []
        by_cat_diff[key].append(q)

    # Répartition cible
    selection_plan = [
        # Déontologie
        ('deontologie', 'facile', 3),
        ('deontologie', 'moyen', 3),
        ('deontologie', 'pointu', 2),
        # Juridique
        ('juridique', 'facile', 2),
        ('juridique', 'moyen', 2),
        ('juridique', 'pointu', 1),
        # Multi-documents (on prendra celles qui ont necessite_multi_documents=true)
        ('multi', 'facile', 1),
        ('multi', 'moyen', 2),
        ('multi', 'pointu', 1),
        # Edge cases
        ('edge', 'facile', 1),
        ('edge', 'moyen', 1),
        ('edge', 'pointu', 1),
    ]

    for category, difficulty, count in selection_plan:
        # Gérer le cas spécial multi-documents
        if category == 'multi':
            # Prendre des questions qui nécessitent plusieurs documents
            candidates = [
                q for q in qa_pairs
                if q.get('necessite_multi_documents', False) and
                q.get('difficulte') == difficulty and
                q not in selected
            ]
        # Gérer le cas spécial edge cases
        elif category == 'edge':
            # Prendre des questions avec catégorie edge ou autre
            candidates = [
                q for q in qa_pairs
                if q.get('categorie', '').lower() in ['edge', 'autre', 'edge_case'] and
                q.get('difficulte') == difficulty and
                q not in selected
            ]
        else:
            # Catégories normales (déontologie, juridique)
            key = f"{category}_{difficulty}"
            candidates = [q for q in by_cat_diff.get(key, []) if q not in selected]

        # Prendre le nombre demandé
        for q in candidates[:count]:
            selected.append(q)
            print(f"  ✓ {category.title()} - {difficulty} : {q.get('question', 'N/A')[:60]}...")

    # Si on n'a pas 20 questions, compléter avec les premières disponibles
    if len(selected) < 20:
        print(f"\n⚠️  Seulement {len(selected)} questions trouvées selon les critères")
        print(f"   Complément avec des questions aléatoires...")

        remaining = [q for q in qa_pairs if q not in selected]
        for q in remaining[:20 - len(selected)]:
            selected.append(q)
            print(f"  ✓ Complément : {q.get('question', 'N/A')[:60]}...")

    print(f"\n✅ {len(selected)} questions sélectionnées")
    return selected[:20]


def generate_excel(
    selected_questions: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path
):
    """
    Génère le fichier Excel pré-rempli avec les 20 questions sélectionnées.

    Args:
        selected_questions: Liste des 20 questions sélectionnées
        template_path: Chemin vers le template Excel
        output_path: Chemin du fichier Excel à créer
    """
    print(f"\n📊 Génération du fichier Excel...")

    # Charger le template
    if not template_path.exists():
        print(f"❌ Template non trouvé : {template_path}")
        print(f"   Exécutez d'abord : python scripts/validation/create_template_validation_dataset.py")
        sys.exit(1)

    wb = load_workbook(template_path)
    ws = wb["Validation_Questions"]

    # Remplir les lignes (à partir de la ligne 2)
    for idx, q in enumerate(selected_questions, start=2):
        # Colonne A : ID
        ws[f'A{idx}'] = q.get('id', f"Q{idx-1:03d}")

        # Colonne B : Question
        ws[f'B{idx}'] = q.get('question', '')

        # Colonne C : Categorie
        ws[f'C{idx}'] = q.get('categorie', '').title()

        # Colonne D : Difficulte
        ws[f'D{idx}'] = q.get('difficulte', '').title()

        # Colonne E : Documents_Sources_Proposes
        sources = q.get('documents_sources_attendus', [])
        ws[f'E{idx}'] = '; '.join(sources) if sources else ''

        # Colonne F : Elements_Cles_Reponse
        elements = q.get('elements_cles_reponse', [])
        if elements:
            # Formater en liste numérotée
            formatted = '\n'.join([f"{i+1}. {elem}" for i, elem in enumerate(elements)])
            ws[f'F{idx}'] = formatted
        else:
            ws[f'F{idx}'] = ''

        # Colonne G : Reponse_Attendue_Resumee
        ws[f'G{idx}'] = q.get('reponse_attendue_resumee', '')

        # Colonnes H à P : Laisser vides (à remplir par l'expert)
        # H: Validation_Question, I: Correction_Question, etc.

        # Ajuster l'alignement
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{idx}'].alignment = Alignment(vertical='top', wrap_text=True)

    # Sauvegarder
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Fichier créé : {output_path}")
    print(f"\n📋 PROCHAINES ÉTAPES :")
    print(f"   1. Ouvrir le fichier Excel : {output_path}")
    print(f"   2. Préparer les liens vers les documents sources")
    print(f"   3. Organiser la session de validation avec les experts (1h30)")
    print(f"   4. Sauvegarder le fichier validé sous : validation_dataset_20questions_VALIDEE.xlsx")


def main():
    """Point d'entrée principal."""
    print("=" * 70)
    print("GÉNÉRATION FICHIER VALIDATION DATASET - PHASE 2")
    print("=" * 70)
    print()

    # Chemins
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    dataset_path = project_root / "tests" / "datasets" / "chatbot_test_dataset.json"
    template_path = project_root / "templates" / "validation_dataset_20questions_TEMPLATE.xlsx"
    output_path = project_root / "output" / "validation_dataset_20questions.xlsx"

    # Vérifications
    if not dataset_path.exists():
        print(f"❌ Fichier dataset introuvable : {dataset_path}")
        sys.exit(1)

    # Charger le dataset
    dataset = load_dataset(dataset_path)

    # Sélectionner 20 questions
    selected = select_20_questions(dataset)

    # Générer l'Excel
    generate_excel(selected, template_path, output_path)

    print()
    print("=" * 70)
    print("✅ GÉNÉRATION TERMINÉE")
    print("=" * 70)


if __name__ == "__main__":
    main()
