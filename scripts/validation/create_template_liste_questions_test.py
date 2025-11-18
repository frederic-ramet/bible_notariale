#!/usr/bin/env python3
"""
Crée le template Excel pour la liste des questions à tester (Phase 3).
Fichier simple utilisé pendant la session de tests du chatbot.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
from pathlib import Path


def create_test_questions_list_template(output_path: str = "templates/liste_questions_a_tester_TEMPLATE.xlsx"):
    """
    Crée un template Excel simple pour la liste des questions à tester.

    Args:
        output_path: Chemin du fichier Excel à créer
    """
    # Créer le classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions_a_Tester"

    # Titre
    ws['A1'] = "LISTE DES 20 QUESTIONS À TESTER - CHATBOT BIBLE NOTARIALE"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 25

    # Instructions
    ws['A2'] = "INSTRUCTIONS : Cocher la case 'Testée' après avoir posé la question au chatbot et donné votre feedback via le système tribunal."
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A2:E2')
    ws.row_dimensions[2].height = 30

    # En-têtes
    headers = ["Numero", "Question", "Categorie", "Testee", "Notes_Rapides"]

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Écrire les en-têtes (ligne 3)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws.row_dimensions[3].height = 30

    # Largeurs de colonnes
    ws.column_dimensions['A'].width = 10  # Numero
    ws.column_dimensions['B'].width = 70  # Question
    ws.column_dimensions['C'].width = 20  # Categorie
    ws.column_dimensions['D'].width = 12  # Testee
    ws.column_dimensions['E'].width = 50  # Notes_Rapides

    # Ajouter 20 lignes vides pour les questions
    for row_idx in range(4, 24):  # Lignes 4 à 23 (20 questions)
        # Numéro
        ws.cell(row=row_idx, column=1, value=row_idx - 3)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")

        # Appliquer les bordures et alignement pour toutes les cellules
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 2:  # Question
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col_idx in [3, 4]:  # Catégorie, Testée
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:  # Notes
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Hauteur de ligne
        ws.row_dimensions[row_idx].height = 40

    # Mise en forme conditionnelle (colonnes "Testée" en jaune pour indiquer qu'elle doit être remplie)
    checkbox_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx in range(4, 24):
        ws.cell(row=row_idx, column=4).fill = checkbox_fill

    # Figer les volets (titres et en-têtes)
    ws.freeze_panes = 'B4'

    # Ajouter un pied de page avec compteur
    ws['A24'] = "TOTAL TESTÉES :"
    ws['A24'].font = Font(bold=True)
    ws['A24'].alignment = Alignment(horizontal="right")
    ws['B24'] = "=COUNTIF(D4:D23,\"x\")"
    ws['B24'].font = Font(bold=True, size=12)
    ws['B24'].alignment = Alignment(horizontal="left")
    ws['C24'] = "/ 20"
    ws['C24'].font = Font(bold=True, size=12)

    # Sauvegarder le fichier
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)
    print(f"✅ Template créé : {output_file}")
    print(f"   - Liste simple de 20 questions")
    print(f"   - Colonne 'Testée' pour cocher les questions testées")
    print(f"   - Colonne 'Notes_Rapides' pour remarques pendant les tests")
    print(f"   - Compteur automatique de questions testées")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "templates/liste_questions_a_tester_TEMPLATE.xlsx"
    create_test_questions_list_template(output_path)
