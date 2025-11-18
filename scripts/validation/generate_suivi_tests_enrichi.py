#!/usr/bin/env python3
"""
Génère le fichier Excel enrichi de suivi des tests pour Phase 3.
Lit le dataset validé et pré-remplit le template avec les questions et réponses de référence.
"""

import json
import sys
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def load_questions(dataset_path: Path) -> List[Dict[str, Any]]:
    """
    Charge les questions depuis le dataset JSON.

    Args:
        dataset_path: Chemin vers le fichier dataset JSON

    Returns:
        Liste des questions
    """
    print(f"📂 Chargement des questions : {dataset_path.name}")

    try:
        with open(dataset_path, 'r', encoding='utf-8') as f:
            dataset = json.load(f)

        # Si c'est le dataset validé (Phase 2), il peut avoir une structure différente
        if 'qa_pairs' in dataset:
            questions = dataset['qa_pairs']
        elif isinstance(dataset, list):
            # Si c'est directement une liste de questions
            questions = dataset
        else:
            print(f"❌ Format de dataset non reconnu")
            sys.exit(1)

        print(f"✅ {len(questions)} questions chargées")
        return questions
    except Exception as e:
        print(f"❌ Erreur lors du chargement : {e}")
        sys.exit(1)


def get_category_label(question: Dict[str, Any]) -> str:
    """Détermine la catégorie d'une question."""
    cat = question.get('categorie', '').lower()

    if 'déonto' in cat or 'deonto' in cat:
        return 'Déontologie'
    elif 'juridique' in cat or 'ccn' in cat or 'rh' in cat:
        return 'Juridique/RH'
    elif 'multi' in cat:
        return 'Multi-documents'
    elif 'edge' in cat or 'hors' in cat:
        return 'Edge case'
    else:
        return cat.title() if cat else 'Autre'


def get_difficulty_label(question: Dict[str, Any]) -> str:
    """Détermine la difficulté d'une question."""
    diff = question.get('difficulte', '').lower()

    if 'facile' in diff:
        return 'Facile'
    elif 'moyen' in diff:
        return 'Moyen'
    elif 'pointu' in diff or 'difficile' in diff:
        return 'Pointu'
    else:
        return diff.title() if diff else ''


def generate_excel(
    questions: List[Dict[str, Any]],
    template_path: Path,
    output_path: Path
):
    """
    Génère le fichier Excel enrichi avec les questions pré-remplies.

    Args:
        questions: Liste des questions
        template_path: Chemin vers le template Excel
        output_path: Chemin du fichier Excel à créer
    """
    print(f"\n📊 Génération du fichier Excel enrichi...")

    # Charger le template
    if not template_path.exists():
        print(f"❌ Template non trouvé : {template_path}")
        print(f"   Exécutez d'abord : python scripts/validation/create_template_suivi_tests_enrichi.py")
        sys.exit(1)

    wb = load_workbook(template_path)
    ws = wb["QA_Tests"]

    # Remplir les lignes (à partir de la ligne 5)
    for idx, q in enumerate(questions, start=5):
        row = idx

        # Colonne A : ID Test (déjà pré-rempli dans le template)
        # Colonne B : Catégorie
        categorie = get_category_label(q)
        ws[f'B{row}'] = categorie
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Colonne C : Question
        question_text = q.get('question', '')
        ws[f'C{row}'] = question_text
        ws[f'C{row}'].alignment = Alignment(vertical='top', wrap_text=True)

        # Colonne D : Document Source
        sources = q.get('documents_sources', [])
        if isinstance(sources, list) and sources:
            # Prendre les 2 premiers documents sources
            doc_sources = ', '.join(sources[:2])
        else:
            doc_sources = str(sources) if sources else ''
        ws[f'D{row}'] = doc_sources
        ws[f'D{row}'].alignment = Alignment(vertical='top', wrap_text=True)

        # Colonne E : Date Test - laisser vide (à remplir pendant le test)

        # Colonne F : Réponse Obtenue - laisser vide (à copier-coller depuis le chatbot)

        # Colonnes G, H, I : Scores - laisser vide (à remplir pendant le test)

        # Colonne J : TOTAL - formule déjà présente dans le template

        # Colonne K : Status - formule déjà présente dans le template

        # Colonne L : Notes - laisser vide

        # Colonne M : Réponse de Référence
        reponse_reference = q.get('reponse_attendue', '')
        if not reponse_reference:
            # Essayer aussi 'reponse' ou 'answer'
            reponse_reference = q.get('reponse', q.get('answer', ''))

        # Si la réponse de référence est trop longue, la résumer
        if isinstance(reponse_reference, str) and len(reponse_reference) > 500:
            # Prendre les éléments clés si disponibles
            elements_cles = q.get('elements_cles_reponse', '')
            if elements_cles:
                reponse_reference = f"Éléments clés attendus :\n{elements_cles}"

        ws[f'M{row}'] = reponse_reference
        ws[f'M{row}'].alignment = Alignment(vertical='top', wrap_text=True)

    # Sauvegarder
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✅ Fichier créé : {output_path}")
    print(f"\n📋 CONTENU :")
    print(f"   - {len(questions)} questions pré-remplies")
    print(f"   - Catégories, Questions, Documents sources")
    print(f"   - Réponses de référence (issues du dataset validé)")
    print(f"\n📝 À REMPLIR PENDANT LES TESTS :")
    print(f"   - Date Test (colonne E)")
    print(f"   - Réponse Obtenue (colonne F - copier-coller depuis le chatbot)")
    print(f"   - Exactitude /3 (colonne G)")
    print(f"   - Sources /3 (colonne H)")
    print(f"   - Formulation /3 (colonne I)")
    print(f"   - Notes/Commentaires (colonne L)")
    print(f"\n⚙️  CALCULS AUTOMATIQUES :")
    print(f"   - TOTAL /9 (colonne J)")
    print(f"   - Status ✅/❌ (colonne K)")
    print(f"   - Onglet 'Synthese' : métriques et décision finale")


def main():
    """Point d'entrée principal."""
    print("=" * 70)
    print("GÉNÉRATION SUIVI ENRICHI DES TESTS - PHASE 3")
    print("=" * 70)
    print()

    # Chemins
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    # Essayer d'abord le dataset validé Phase 2, sinon le dataset original
    dataset_validated_path = project_root / "tests" / "datasets" / "dataset_test_final_20questions.json"
    dataset_original_path = project_root / "tests" / "datasets" / "chatbot_test_dataset.json"

    # Déterminer quel dataset utiliser
    if dataset_validated_path.exists():
        print(f"✅ Utilisation du dataset validé Phase 2 : {dataset_validated_path.name}")
        dataset_path = dataset_validated_path
    else:
        print(f"⚠️  Dataset validé JSON non trouvé")
        print(f"   Utilisation du dataset original : {dataset_original_path.name}")
        dataset_path = dataset_original_path

    template_path = project_root / "templates" / "suivi_tests_chatbot_TEMPLATE.xlsx"
    output_path = project_root / "output" / "suivi_tests_chatbot.xlsx"

    # Vérifications
    if not dataset_path.exists():
        print(f"❌ Aucun fichier dataset trouvé")
        sys.exit(1)

    # Charger les questions
    questions = load_questions(dataset_path)

    # Limiter à 20 questions max
    if len(questions) > 20:
        print(f"ℹ️  {len(questions)} questions trouvées, on prend les 20 premières")
        questions = questions[:20]

    # Générer l'Excel
    generate_excel(questions, template_path, output_path)

    print()
    print("=" * 70)
    print("✅ GÉNÉRATION TERMINÉE")
    print("=" * 70)
    print()
    print("💡 UTILISATION :")
    print(f"   1. Ouvrir : {output_path}")
    print(f"   2. Pour chaque question :")
    print(f"      - Poser la question au chatbot")
    print(f"      - Copier-coller la réponse dans 'Réponse Obtenue'")
    print(f"      - Noter sur les 3 critères (/3 chacun)")
    print(f"      - Ajouter des notes/commentaires si besoin")
    print(f"   3. Consulter l'onglet 'Synthese' pour les métriques")
    print(f"   4. Décision finale calculée automatiquement")


if __name__ == "__main__":
    main()
