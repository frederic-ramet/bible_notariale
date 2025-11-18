#!/usr/bin/env python3
"""
Crée le template Excel ENRICHI pour le suivi des tests du chatbot (Phase 3).
Inspiré du fichier BM_QA_Marianne_test20250611_BM.xlsx mais avec notre système /9.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path


def create_enriched_test_tracking_template(output_path: str = "templates/suivi_tests_chatbot_TEMPLATE.xlsx"):
    """
    Crée un template Excel enrichi pour le suivi détaillé des tests du chatbot.

    Structure :
    - Onglet 1 : QA_Tests (suivi détaillé de chaque test)
    - Onglet 2 : Synthese (métriques automatiques)

    Args:
        output_path: Chemin du fichier Excel à créer
    """
    wb = Workbook()

    # ========================================================================
    # ONGLET 1 : QA_Tests
    # ========================================================================
    ws_qa = wb.active
    ws_qa.title = "QA_Tests"

    # Titre principal
    ws_qa['A1'] = "SUIVI DES TESTS CHATBOT - BIBLE NOTARIALE"
    ws_qa['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_qa['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_qa['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_qa.merge_cells('A1:M1')
    ws_qa.row_dimensions[1].height = 30

    # Sous-titre avec instructions
    ws_qa['A2'] = "Système d'évaluation : Exactitude /3 + Sources /3 + Formulation /3 = TOTAL /9 (Seuil de réussite : ≥6/9)"
    ws_qa['A2'].font = Font(italic=True, size=10, color="666666")
    ws_qa['A2'].alignment = Alignment(horizontal="center", wrap_text=True)
    ws_qa.merge_cells('A2:M2')
    ws_qa.row_dimensions[2].height = 25

    # En-têtes principales (ligne 3)
    headers_main = [
        "ID Test",
        "Catégorie",
        "Question",
        "Document Source",
        "Date Test",
        "Réponse Obtenue",
        "Exactitude /3",
        "Sources /3",
        "Formulation /3",
        "TOTAL /9",
        "Status",
        "Notes",
        "Réponse de Référence"
    ]

    # Sous-en-têtes explicatifs (ligne 4)
    headers_sub = [
        "",
        "",
        "",
        "",
        "",
        "",
        "Pertinence + Complétude",
        "Pertinence + Complétude",
        "Clarté + Style + Longueur",
        "",
        "✅ si ≥6, ❌ si <6",
        "Commentaires libres",
        "Issue du dataset validé"
    ]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(italic=True, size=9, color="FFFFFF")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Écrire les en-têtes principaux (ligne 3)
    for col_idx, header in enumerate(headers_main, 1):
        cell = ws_qa.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    ws_qa.row_dimensions[3].height = 35

    # Écrire les sous-en-têtes (ligne 4)
    for col_idx, subheader in enumerate(headers_sub, 1):
        cell = ws_qa.cell(row=4, column=col_idx, value=subheader)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = header_alignment
        cell.border = border

    ws_qa.row_dimensions[4].height = 30

    # Largeurs de colonnes
    column_widths = {
        'A': 15,   # ID Test
        'B': 15,   # Catégorie
        'C': 50,   # Question
        'D': 20,   # Document Source
        'E': 12,   # Date Test
        'F': 60,   # Réponse Obtenue
        'G': 12,   # Exactitude
        'H': 12,   # Sources
        'I': 12,   # Formulation
        'J': 12,   # TOTAL
        'K': 10,   # Status
        'L': 40,   # Notes
        'M': 60    # Réponse de Référence
    }

    for col_letter, width in column_widths.items():
        ws_qa.column_dimensions[col_letter].width = width

    # Ajouter 20 lignes pour les tests (lignes 5 à 24)
    score_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for row_idx in range(5, 25):  # 20 questions
        # ID Test
        test_id = f"TEST_{row_idx - 4:03d}"
        ws_qa.cell(row=row_idx, column=1, value=test_id)
        ws_qa.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")

        # Date Test - formule automatique pour date du jour
        ws_qa.cell(row=row_idx, column=5, value="")  # Vide, à remplir manuellement
        ws_qa.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="top")

        # Colonnes de scores - fond jaune pour indiquer qu'elles doivent être remplies
        for score_col in [7, 8, 9]:  # G, H, I
            ws_qa.cell(row=row_idx, column=score_col).fill = score_fill
            ws_qa.cell(row=row_idx, column=score_col).alignment = Alignment(horizontal="center", vertical="top")

        # TOTAL - formule automatique
        total_formula = f"=SUM(G{row_idx}:I{row_idx})"
        ws_qa.cell(row=row_idx, column=10, value=total_formula)
        ws_qa.cell(row=row_idx, column=10).alignment = Alignment(horizontal="center", vertical="top")
        ws_qa.cell(row=row_idx, column=10).font = Font(bold=True)

        # Status - formule conditionnelle
        status_formula = f'=IF(J{row_idx}="","",IF(J{row_idx}>=6,"✅ Réussi","❌ Échec"))'
        ws_qa.cell(row=row_idx, column=11, value=status_formula)
        ws_qa.cell(row=row_idx, column=11).alignment = Alignment(horizontal="center", vertical="top")

        # Appliquer les bordures et alignement
        for col_idx in range(1, 14):
            cell = ws_qa.cell(row=row_idx, column=col_idx)
            cell.border = border

            # Alignement selon la colonne
            if col_idx in [1, 2, 5, 7, 8, 9, 10, 11]:  # Colonnes centrées
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif col_idx in [3, 4, 6, 12, 13]:  # Colonnes avec texte long
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Hauteur de ligne
        ws_qa.row_dimensions[row_idx].height = 60

    # Figer les volets (2 lignes de titre + 2 lignes d'en-têtes)
    ws_qa.freeze_panes = 'D5'

    # ========================================================================
    # ONGLET 2 : Synthese
    # ========================================================================
    ws_synth = wb.create_sheet("Synthese")

    # Titre
    ws_synth['A1'] = "📈 SYNTHÈSE DES TESTS"
    ws_synth['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_synth['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_synth['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_synth.merge_cells('A1:C1')
    ws_synth.row_dimensions[1].height = 30

    # En-têtes du tableau de synthèse
    ws_synth['A3'] = "Métrique"
    ws_synth['B3'] = "Valeur"
    ws_synth['C3'] = "Objectif"

    for col in ['A', 'B', 'C']:
        ws_synth[f'{col}3'].font = Font(bold=True, size=11)
        ws_synth[f'{col}3'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws_synth[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws_synth[f'{col}3'].alignment = Alignment(horizontal="center", vertical="center")
        ws_synth[f'{col}3'].border = border

    # Métriques
    metrics = [
        ("Total Tests", "=COUNTA(QA_Tests!A5:A24)", "20"),
        ("Tests Exécutés", '=COUNTIF(QA_Tests!K5:K24,"✅ Réussi")+COUNTIF(QA_Tests!K5:K24,"❌ Échec")', "20"),
        ("Réussis (≥6/9)", '=COUNTIF(QA_Tests!K5:K24,"✅ Réussi")', "≥16 (80%)"),
        ("Échecs (<6/9)", '=COUNTIF(QA_Tests!K5:K24,"❌ Échec")', "≤4 (20%)"),
        ("Score Moyen", "=AVERAGE(QA_Tests!J5:J24)", "≥6/9"),
        ("% Réussite", "=IF(B5=0,0,B6/B5*100)", "≥80%"),
        ("", "", ""),
        ("Score Moyen Exactitude", "=AVERAGE(QA_Tests!G5:G24)", "/3"),
        ("Score Moyen Sources", "=AVERAGE(QA_Tests!H5:H24)", "/3"),
        ("Score Moyen Formulation", "=AVERAGE(QA_Tests!I5:I24)", "/3"),
    ]

    row_start = 4
    for idx, (metric, formula, target) in enumerate(metrics, row_start):
        ws_synth[f'A{idx}'] = metric
        ws_synth[f'B{idx}'] = formula if formula else ""
        ws_synth[f'C{idx}'] = target

        # Style
        if metric == "":  # Ligne vide
            continue

        ws_synth[f'A{idx}'].border = border
        ws_synth[f'B{idx}'].border = border
        ws_synth[f'C{idx}'].border = border

        ws_synth[f'A{idx}'].font = Font(bold=True if idx <= 6 else False)
        ws_synth[f'B{idx}'].alignment = Alignment(horizontal="center")
        ws_synth[f'C{idx}'].alignment = Alignment(horizontal="center")

        # Format pourcentage pour % Réussite
        if metric == "% Réussite":
            ws_synth[f'B{idx}'].number_format = '0.0"%"'
        # Format décimal pour les moyennes
        elif "Moyen" in metric:
            ws_synth[f'B{idx}'].number_format = '0.0'

    # Largeurs de colonnes
    ws_synth.column_dimensions['A'].width = 30
    ws_synth.column_dimensions['B'].width = 15
    ws_synth.column_dimensions['C'].width = 20

    # Décision finale
    ws_synth['A16'] = "DÉCISION FINALE"
    ws_synth['A16'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_synth['A16'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_synth['A16'].alignment = Alignment(horizontal="center", vertical="center")
    ws_synth.merge_cells('A16:C16')
    ws_synth.row_dimensions[16].height = 25

    decision_text = """=IF(B7>=80%,"✅ GO PHASE 2 : Déploiement élargi",IF(B7>=60%,"⚠️ ITÉRATION : Corrections ciblées + re-tests","❌ STOP : Revoir l'architecture"))"""

    ws_synth['A17'] = decision_text
    ws_synth['A17'].font = Font(size=12, bold=True)
    ws_synth['A17'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_synth.merge_cells('A17:C17')
    ws_synth.row_dimensions[17].height = 40

    # Sauvegarder le fichier
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)

    print(f"✅ Template enrichi créé : {output_file}")
    print(f"   - Onglet 'QA_Tests' : Suivi détaillé de chaque test")
    print(f"   - Onglet 'Synthese' : Métriques automatiques")
    print(f"   - Système d'évaluation : /9 (Exactitude /3 + Sources /3 + Formulation /3)")
    print(f"   - Colonnes : Réponse Obtenue + Réponse de Référence")
    print(f"   - Status automatique : ✅ si ≥6/9, ❌ si <6/9")
    print(f"   - Décision finale calculée automatiquement")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "templates/suivi_tests_chatbot_TEMPLATE.xlsx"
    create_enriched_test_tracking_template(output_path)
