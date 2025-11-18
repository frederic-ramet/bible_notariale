#!/usr/bin/env python3
"""
Crée le template Excel pour la validation des métadonnées (20 documents).
Ce fichier sera utilisé par l'expert métier lors de la session de validation Phase 1.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import sys
from pathlib import Path


def create_metadata_validation_template(output_path: str = "templates/validation_metadonnees_20docs_TEMPLATE.xlsx"):
    """
    Crée un template Excel pour la validation des métadonnées.

    Args:
        output_path: Chemin du fichier Excel à créer
    """
    # Créer le classeur
    wb = Workbook()

    # ===== ONGLET 1 : INSTRUCTIONS =====
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"

    # Titre
    ws_instructions['A1'] = "GUIDE DE VALIDATION DES MÉTADONNÉES"
    ws_instructions['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_instructions['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_instructions.merge_cells('A1:E1')
    ws_instructions.row_dimensions[1].height = 30

    # Contenu des instructions
    instructions = [
        ("", ""),
        ("OBJECTIF", "Valider que l'annotation automatique des 20 documents est correcte"),
        ("DURÉE", "2 heures (environ 5-6 minutes par document)"),
        ("", ""),
        ("COMMENT UTILISER CE FICHIER", ""),
        ("", ""),
        ("1. Pour chaque document :", "Consulter le PDF source"),
        ("2. Valider le TYPE", "Sélectionner 'OK' ou 'A corriger' dans la liste déroulante"),
        ("", "Si 'A corriger' : Indiquer le bon type dans la colonne suivante"),
        ("3. Valider les CATÉGORIES", "Sélectionner 'OK' ou 'A corriger'"),
        ("", "Si 'A corriger' : Indiquer les bonnes catégories (séparées par des virgules)"),
        ("4. Valider la PRIORITÉ", "Sélectionner 'OK' ou 'A corriger'"),
        ("", "Si 'A corriger' : Indiquer le bon niveau (1-10)"),
        ("5. Commentaires", "Ajouter des remarques libres si nécessaire"),
        ("", ""),
        ("ÉCHELLE DE PRIORITÉ", ""),
        ("10", "Documents critiques (RPN, Code déontologie, Circulaires majeures)"),
        ("8-9", "Documents importants (Guides CSN, Avenants CCN majeurs)"),
        ("5-7", "Documents utiles (Fil-Infos récents, Guides pratiques)"),
        ("1-4", "Documents secondaires"),
        ("", ""),
        ("CATÉGORIES MÉTIER DISPONIBLES", ""),
        ("", "DEONTOLOGIE, RH, FORMATION, NEGOCIATION_IMMOBILIERE, PROCEDURE,"),
        ("", "TARIFICATION, LCB_FT, ASSURANCE, ORGANISATION_PROFESSION, AUTRE"),
        ("", ""),
        ("TYPES DE DOCUMENTS POSSIBLES", ""),
        ("", "Circulaire CSN, Règlement professionnel, Code de déontologie, Avenant CCN,"),
        ("", "Accord de branche, Fil-Info, Guide pratique, Ordonnance, Décret, Arrêté, Autre"),
    ]

    row = 2
    for col1, col2 in instructions:
        if col1.isupper() and col1:  # Titres de section
            ws_instructions[f'A{row}'] = col1
            ws_instructions[f'A{row}'].font = Font(bold=True, size=12)
            ws_instructions[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            if col2:
                ws_instructions[f'B{row}'] = col2
                ws_instructions[f'B{row}'].font = Font(bold=True)
        else:
            ws_instructions[f'A{row}'] = col1
            ws_instructions[f'B{row}'] = col2
        row += 1

    # Ajuster les largeurs de colonnes
    ws_instructions.column_dimensions['A'].width = 35
    ws_instructions.column_dimensions['B'].width = 70

    # ===== ONGLET 2 : VALIDATION_METADONNEES =====
    ws_validation = wb.create_sheet("Validation_Metadonnees")

    # En-têtes
    headers = [
        "ID",
        "Nom_Fichier",
        "Type_Propose",
        "Categories_Proposees",
        "Priorite_Proposee",
        "Mots_Cles_Proposes",
        "Validation_Type",
        "Correction_Type",
        "Validation_Categories",
        "Correction_Categories",
        "Validation_Priorite",
        "Correction_Priorite",
        "Commentaires"
    ]

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws_validation.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Hauteur de la ligne d'en-tête
    ws_validation.row_dimensions[1].height = 40

    # Largeurs de colonnes
    column_widths = {
        'A': 25,  # ID
        'B': 40,  # Nom_Fichier
        'C': 25,  # Type_Propose
        'D': 35,  # Categories_Proposees
        'E': 12,  # Priorite_Proposee
        'F': 50,  # Mots_Cles_Proposes
        'G': 15,  # Validation_Type
        'H': 30,  # Correction_Type
        'I': 20,  # Validation_Categories
        'J': 35,  # Correction_Categories
        'K': 18,  # Validation_Priorite
        'L': 18,  # Correction_Priorite
        'M': 50,  # Commentaires
    }

    for col, width in column_widths.items():
        ws_validation.column_dimensions[col].width = width

    # Ajouter 20 lignes vides pour les données
    for row_idx in range(2, 22):  # Lignes 2 à 21 (20 documents)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_validation.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ===== LISTES DÉROULANTES =====

    # Liste déroulante pour Validation_Type (colonne G)
    dv_type = DataValidation(
        type="list",
        formula1='"OK,A corriger"',
        allow_blank=True
    )
    dv_type.error = 'Veuillez choisir une valeur dans la liste'
    dv_type.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_type)
    dv_type.add(f'G2:G21')

    # Liste déroulante pour Validation_Categories (colonne I)
    dv_cat = DataValidation(
        type="list",
        formula1='"OK,A corriger"',
        allow_blank=True
    )
    dv_cat.error = 'Veuillez choisir une valeur dans la liste'
    dv_cat.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_cat)
    dv_cat.add(f'I2:I21')

    # Liste déroulante pour Validation_Priorite (colonne K)
    dv_prio = DataValidation(
        type="list",
        formula1='"OK,A corriger"',
        allow_blank=True
    )
    dv_prio.error = 'Veuillez choisir une valeur dans la liste'
    dv_prio.errorTitle = 'Valeur invalide'
    ws_validation.add_data_validation(dv_prio)
    dv_prio.add(f'K2:K21')

    # Mise en forme conditionnelle (simulation avec couleurs de fond)
    # Les colonnes de validation seront en jaune pâle pour indiquer qu'elles doivent être remplies
    validation_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row_idx in range(2, 22):
        ws_validation[f'G{row_idx}'].fill = validation_fill
        ws_validation[f'I{row_idx}'].fill = validation_fill
        ws_validation[f'K{row_idx}'].fill = validation_fill

    # Protection des colonnes de données (optionnel - commenté pour permettre l'édition)
    # On pourrait verrouiller les colonnes A-F pour éviter les modifications accidentelles

    # Figer les volets (première ligne et première colonne)
    ws_validation.freeze_panes = 'B2'

    # Sauvegarder le fichier
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_file)
    print(f"✅ Template créé : {output_file}")
    print(f"   - Onglet 'Instructions' : Guide complet pour l'expert")
    print(f"   - Onglet 'Validation_Metadonnees' : 20 lignes prêtes à remplir")
    print(f"   - Listes déroulantes configurées pour les colonnes de validation")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "templates/validation_metadonnees_20docs_TEMPLATE.xlsx"
    create_metadata_validation_template(output_path)
