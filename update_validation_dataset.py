#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour mettre à jour le fichier Excel validation_dataset_20questions.xlsx
Ajoute des onglets de référence et des listes déroulantes
"""

import json
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# Configuration
BASE_DIR = Path(__file__).parent
METADATA_DIR = BASE_DIR / "_metadata"
INDEX_FILE = METADATA_DIR / "index_complet.json"
EXCEL_FILE = BASE_DIR / "output" / "validation_dataset_20questions.xlsx"
OUTPUT_FILE = BASE_DIR / "output" / "validation_dataset_20questions_updated.xlsx"


def extract_reference_data():
    """
    Extrait les données de référence depuis l'index complet
    """
    with open(INDEX_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    documents = data.get('documents', [])

    # 1. Types de document
    types_doc = set()
    for doc in documents:
        type_doc = doc.get('classification', {}).get('type_document', '')
        if type_doc:
            types_doc.add(type_doc)

    # 2. Thématiques
    thematiques = set()
    for doc in documents:
        themes = doc.get('classification', {}).get('thematiques', [])
        for theme in themes:
            if theme:
                thematiques.add(theme)

    # 3. Mots-clés
    mots_cles = set()
    for doc in documents:
        mots = doc.get('mots_cles', [])
        for mot in mots:
            if mot:
                mots_cles.add(mot)

    # 4. Sources de document
    sources = set()
    for doc in documents:
        source = doc.get('classification', {}).get('sources_document', '')
        if source:
            sources.add(source)

    # 5. Domaines métier
    domaines = set()
    for doc in documents:
        doms = doc.get('classification', {}).get('domaines_metier', [])
        for dom in doms:
            if dom:
                domaines.add(dom)

    return {
        'types_document': sorted(types_doc),
        'thematiques': sorted(thematiques),
        'mots_cles': sorted(mots_cles),
        'sources_document': sorted(sources),
        'domaines_metier': sorted(domaines)
    }


def create_reference_sheets(wb, ref_data):
    """
    Crée les onglets de référence avec les valeurs
    """
    # Supprimer les onglets de référence existants s'ils existent
    for sheet_name in ['Ref_TypesDocument', 'Ref_Thematiques', 'Ref_MotsCles',
                       'Ref_SourcesDocument', 'Ref_DomainesMetier']:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # 1. Onglet Types de document
    ws_types = wb.create_sheet('Ref_TypesDocument')
    ws_types.append(['Type de document'])
    for type_doc in ref_data['types_document']:
        ws_types.append([type_doc])

    # 2. Onglet Thématiques
    ws_themes = wb.create_sheet('Ref_Thematiques')
    ws_themes.append(['Thématique'])
    for theme in ref_data['thematiques']:
        ws_themes.append([theme])

    # 3. Onglet Mots-clés
    ws_mots = wb.create_sheet('Ref_MotsCles')
    ws_mots.append(['Mot-clé'])
    for mot in ref_data['mots_cles']:
        ws_mots.append([mot])

    # 4. Onglet Sources de document
    ws_sources = wb.create_sheet('Ref_SourcesDocument')
    ws_sources.append(['Source de document'])
    for source in ref_data['sources_document']:
        ws_sources.append([source])

    # 5. Onglet Domaines métier
    ws_domaines = wb.create_sheet('Ref_DomainesMetier')
    ws_domaines.append(['Domaine métier'])
    for domaine in ref_data['domaines_metier']:
        ws_domaines.append([domaine])

    print("✅ Onglets de référence créés")


def add_columns_with_dropdowns(wb, ref_data):
    """
    Ajoute des colonnes avec listes déroulantes dans l'onglet Validation_Questions
    """
    ws = wb['Validation_Questions']

    # Position d'insertion : après la colonne "Categorie" (col 3)
    insert_after_col = 3

    # Insérer 5 nouvelles colonnes
    ws.insert_cols(insert_after_col + 1, 5)

    # Ajouter les en-têtes
    ws.cell(1, insert_after_col + 1).value = 'Type_Document'
    ws.cell(1, insert_after_col + 2).value = 'Domaine_Metier'
    ws.cell(1, insert_after_col + 3).value = 'Source_Document'
    ws.cell(1, insert_after_col + 4).value = 'Thematiques'
    ws.cell(1, insert_after_col + 5).value = 'Mots_Cles'

    # Nombre de lignes de données (exclure l'en-tête)
    max_row = ws.max_row

    # Créer les listes déroulantes
    # 1. Type de document
    dv_type = DataValidation(
        type="list",
        formula1=f"=Ref_TypesDocument!$A$2:$A${len(ref_data['types_document']) + 1}",
        allow_blank=True
    )
    dv_type.prompt = "Choisissez un type de document"
    dv_type.promptTitle = "Type de document"
    ws.add_data_validation(dv_type)
    for row in range(2, max_row + 1):
        dv_type.add(ws.cell(row, insert_after_col + 1))

    # 2. Domaine métier
    dv_domaine = DataValidation(
        type="list",
        formula1=f"=Ref_DomainesMetier!$A$2:$A${len(ref_data['domaines_metier']) + 1}",
        allow_blank=True
    )
    dv_domaine.prompt = "Choisissez un domaine métier"
    dv_domaine.promptTitle = "Domaine métier"
    ws.add_data_validation(dv_domaine)
    for row in range(2, max_row + 1):
        dv_domaine.add(ws.cell(row, insert_after_col + 2))

    # 3. Source de document
    dv_source = DataValidation(
        type="list",
        formula1=f"=Ref_SourcesDocument!$A$2:$A${len(ref_data['sources_document']) + 1}",
        allow_blank=True
    )
    dv_source.prompt = "Choisissez une source de document"
    dv_source.promptTitle = "Source de document"
    ws.add_data_validation(dv_source)
    for row in range(2, max_row + 1):
        dv_source.add(ws.cell(row, insert_after_col + 3))

    # 4. Thématiques (liste longue, permet saisie personnalisée)
    dv_theme = DataValidation(
        type="list",
        formula1=f"=Ref_Thematiques!$A$2:$A${len(ref_data['thematiques']) + 1}",
        allow_blank=True
    )
    dv_theme.prompt = "Choisissez une ou plusieurs thématiques (séparées par des virgules)"
    dv_theme.promptTitle = "Thématiques"
    ws.add_data_validation(dv_theme)
    for row in range(2, max_row + 1):
        dv_theme.add(ws.cell(row, insert_after_col + 4))

    # 5. Mots-clés (liste longue, permet saisie personnalisée)
    dv_mots = DataValidation(
        type="list",
        formula1=f"=Ref_MotsCles!$A$2:$A${len(ref_data['mots_cles']) + 1}",
        allow_blank=True
    )
    dv_mots.prompt = "Choisissez un ou plusieurs mots-clés (séparés par des virgules)"
    dv_mots.promptTitle = "Mots-clés"
    ws.add_data_validation(dv_mots)
    for row in range(2, max_row + 1):
        dv_mots.add(ws.cell(row, insert_after_col + 5))

    print("✅ Colonnes et listes déroulantes ajoutées")


def main():
    """
    Fonction principale
    """
    print("\n🚀 Mise à jour du fichier Excel validation_dataset_20questions.xlsx\n")

    # 1. Extraire les données de référence
    print("📊 Extraction des données de référence depuis l'index...")
    ref_data = extract_reference_data()

    print(f"  - Types de document: {len(ref_data['types_document'])}")
    print(f"  - Domaines métier: {len(ref_data['domaines_metier'])}")
    print(f"  - Sources de document: {len(ref_data['sources_document'])}")
    print(f"  - Thématiques: {len(ref_data['thematiques'])}")
    print(f"  - Mots-clés: {len(ref_data['mots_cles'])}")

    # 2. Charger le fichier Excel
    print(f"\n📂 Chargement du fichier Excel : {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # 3. Créer les onglets de référence
    print("\n📋 Création des onglets de référence...")
    create_reference_sheets(wb, ref_data)

    # 4. Ajouter les colonnes avec listes déroulantes
    print("\n🔽 Ajout des colonnes avec listes déroulantes...")
    add_columns_with_dropdowns(wb, ref_data)

    # 5. Sauvegarder le fichier mis à jour
    print(f"\n💾 Sauvegarde du fichier mis à jour : {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    wb.close()

    print("\n" + "="*80)
    print("✅ MISE À JOUR TERMINÉE AVEC SUCCÈS")
    print("="*80)
    print(f"\n📁 Fichier de sortie : {OUTPUT_FILE}")
    print("\n📌 Nouvelles colonnes ajoutées :")
    print("  1. Type_Document (liste déroulante)")
    print("  2. Domaine_Metier (liste déroulante)")
    print("  3. Source_Document (liste déroulante)")
    print("  4. Thematiques (liste déroulante)")
    print("  5. Mots_Cles (liste déroulante)")
    print("\n📚 Onglets de référence créés :")
    print("  - Ref_TypesDocument")
    print("  - Ref_DomainesMetier")
    print("  - Ref_SourcesDocument")
    print("  - Ref_Thematiques")
    print("  - Ref_MotsCles")
    print()


if __name__ == '__main__':
    main()
